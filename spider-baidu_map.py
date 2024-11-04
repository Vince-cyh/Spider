import json
import requests
import openpyxl

###请求头
headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/89.0.4389.90 Safari/537.36",
        'Referer':'https://map.baidu.com/search/%E7%B2%A4%E8%8F%9C/@11606355.219793778,4669275.879793777,5z',
        "Cookie":"BIDUPSID=1D2214688117613CC11AC890AC7CB266; PSTM=1668828964; BAIDUID_BFESS=1D2214688117613CC11AC890AC7CB266:FG=1; __bid_n=18871a59b8b2656c571d88; ZFY=:AAzQIF03p:ANNm8Gg02:Bdrl6cGCzoZSb:BYRoc6QWIOyE:C; FPTOKEN=/mf6vvRzn9LY3wV2vuBufbZniuUJFEc0B+vRsFOjAHc+FDJfDXChAGYbNlRdeKW0YYXo/Z3DyvBJYSoLcNHFBp37GiHAeV6YJITiUTVe4Zo7aXUmAcW6ILHJlF5iZM9jqQ1Ri783a92gHbVSC4sALNkPdIxlOm2NPjxSGYsxz3ogkCK1XEAqbPlXE292wITXaOdQfkJLveG/Tl0PxOjJI0CEvUuyk+qhDfdOcE6YCPjpHIZdKiMu9XFftJAw3zc9D7+KH7ie4OYiOg85cgtF+a6dk2uB59oAiWrEfT66BkULqDxqUqkwKbVShkIxgZ3eF+HAqvI0NDxHs/sFZ9Jc02lvUhEJS+5bE1xRix5PhIgMKH6PoSxOcDiu9Q6oIG1I+oydqE8aLs2WHAt452DcAg==|S7K9B5krexDlDDhVhqbBgvcccPRnUEzRCVIskAlBm2Y=|10|4793e06429a4bc5574068d165c56c66c; newlogin=1; BDUSS=NuRWdUN25BQTdaalpBejhST2pvZDNUaE1IRng2Z2x6Zk04a1FPVjc3ZmlaQkpsRVFBQUFBJCQAAAAAAAAAAAEAAABQdPrnSGdWdnZpbmNlAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAOLX6mTi1-pkcz; BDUSS_BFESS=NuRWdUN25BQTdaalpBejhST2pvZDNUaE1IRng2Z2x6Zk04a1FPVjc3ZmlaQkpsRVFBQUFBJCQAAAAAAAAAAAEAAABQdPrnSGdWdnZpbmNlAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAOLX6mTi1-pkcz; MCITY=-%3A; M_LG_UID=3891950672; M_LG_SALT=50db37554b584f0858a6b3e0f464d913; ab_sr=1.0.1_MDkyYzIzYjUxNTkyYzI2NGI3Y2EwMTI1YmEwMGQyZTMyMTM2Y2UxNDA2YTBkNjk0NzM0N2E1OGNmMDY3ODNlMDliYTVhYmE5MWJhNDc0OGQyNGY1ZGZjMDY4MDAxYmNhNjZmOThiZjU3YTNlMWJkODI2NjUxNzdlNzlmZDFmOWFmZmY0MTdlZjgyM2M0YjViZTExZDEyMTgyYmFkYzA1MWJkMTdjNDVjZWUyNzVjYjVhZmQ4MzhkMWJiYjVkYTc5; validate=59007"
}
##请求链接
url = "https://map.baidu.com/?newmap=1&reqflag=pcmap&biz=1&from=webmap&da_par=direct&pcevaname=pc4.1&qt=s&da_src=shareurl&wd=%E7%B2%A4%E8%8F%9C&c=1&src=0&pn=0&sug=0&l=5&b=(9206099.22,1740635.88;14006611.22,7597915.88)&from=webmap&biz_forward=%7B%22scaler%22:2,%22styles%22:%22pl%22%7D&device_ratio=2&auth=GU0RxQOd%40H1OBE7%40aLvE8cTFYd8xM%40AVuxLTHENzNEztAXA7271XA5xAwwyS8v7uvkGcuVtvvhguVtvyheuVtvCMGuVtvCQMuVtvIPcuVtvYvjuVtvZgMuVtv%40vcuVtvc3CuVtvcPPuVtveGvuVtveh3uVtvh3CuVtvhgMuxVVtvrMhuxtdw8E62qvyUuUuuouKi3egvcguxLTHENzNEz&seckey=nIKpwi%2BsaGwBYM%2FBfFBhQgizgJZ8QA4wKxt90g%2FnUSw%3D%2C7vTEYNhMpD6e_kEFdKdpq6DiqX9S8YqfjeAea9hVyXl2aROK34w4YS9QaMOMf5EbQJstJ0J8erXNh_L-GM5bleRDQWCz29I-89Mae-rQGVN2vWYynm1Mqaat7UsnIOEpii_Qb76FtMOLYyr3gBj092oq_yyvi6B-YJyiu8TWEj6Wfl38PsHzU6b6f3-vwvZO&tn=B_NORMAL_MAP&nn=0&u_loc=12683830,2559879&ie=utf-8&t=1695472743192&newfrom=zhuzhan_webmap"

outwb_p = openpyxl.Workbook()
outws_p = outwb_p.create_sheet(index=0)
outws_p.cell(row=1, column=1, value="省份")
outws_p.cell(row=1, column=2, value="数量")

outwb_c = openpyxl.Workbook()
outws_c = outwb_c.create_sheet(index=0)
outws_c.cell(row=1, column=1, value="城市")
outws_c.cell(row=1, column=2, value="数量")

###响应数据

hot_city = requests.get(url,headers=headers).json()
hot_city = hot_city['content']

t = 1
p = 1
for i in hot_city:
    t = t + 1
    outws_c.cell(row=t, column=1, value=i['name'])
    outws_c.cell(row=t, column=2, value=i['num'])

response = requests.get(url,headers=headers).json()
response = response['more_city']

t = 7
p = 7

for i in response:
    t = t + 1
    city = i['city']
    print(i['province'])
    print(i['num'])
    outws_p.cell(row=t, column=1, value=i['province'])
    outws_p.cell(row=t, column=2, value=i['num'])
    for j in city:
        p = p + 1
        print(j['name'])
        print(j['num'])
        outws_c.cell(row=p, column=1, value=j['name'])
        outws_c.cell(row=p, column=2, value=j['num'])

outwb_p.save("全国省份粤菜数量.xlsx")  # 保存
outwb_c.save("全国城市粤菜数量.xlsx")  # 保存

